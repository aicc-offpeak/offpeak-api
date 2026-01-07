#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build zones seed JSON from Seoul 120/82 locations XLSX.

Usage:
  1) Download XLSX:
     - "서울시 주요 120장소 목록.xlsx" (Seoul Open Data Plaza)
     - "서울시 주요 82장소 목록.xlsx" (Seoul Open Data Plaza)

  2) Put it anywhere, then run:
     python scripts/build_zones_seed.py \
       --input "/path/to/서울시 주요 120장소 목록.xlsx" \
       --output "app/resources/zones_seed.json" \
       --geocode

Notes:
  - If the XLSX doesn't contain lat/lng, use --geocode to fill coordinates via Kakao keyword search.
  - Requires:
      pip install openpyxl httpx
  - Env:
      KAKAO_REST_API_KEY=xxxxxxxxxxxxxxxx
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import httpx
from openpyxl import load_workbook


# -----------------------------
# Kakao keyword search (geocode)
# -----------------------------
KAKAO_BASE_URL = "https://dapi.kakao.com"
SEOUL_CENTER_LAT = 37.5665
SEOUL_CENTER_LNG = 126.9780


def kakao_keyword_geocode(
    *,
    query: str,
    rest_api_key: str,
    timeout_s: float = 6.0,
    bias_center: Tuple[float, float] = (SEOUL_CENTER_LAT, SEOUL_CENTER_LNG),
    radius_m: int = 20000,
) -> Optional[Tuple[float, float]]:
    """
    Returns (lat, lng) from Kakao keyword search. None if not found.
    """
    url = f"{KAKAO_BASE_URL}/v2/local/search/keyword.json"
    headers = {"Authorization": f"KakaoAK {rest_api_key}"}
    params = {
        "query": query,
        "x": f"{bias_center[1]:.7f}",  # lng
        "y": f"{bias_center[0]:.7f}",  # lat
        "radius": int(max(0, min(radius_m, 20000))),
        "size": 1,
        "sort": "distance",
    }

    with httpx.Client(timeout=timeout_s) as client:
        r = client.get(url, headers=headers, params=params)

    if r.status_code != 200:
        raise RuntimeError(f"Kakao keyword search failed: {r.status_code} {r.text}")

    data = r.json()
    docs = data.get("documents") or []
    if not docs:
        return None

    top = docs[0]
    lat = _safe_float(top.get("y"))
    lng = _safe_float(top.get("x"))
    if lat == 0.0 and lng == 0.0:
        return None
    return (lat, lng)


# -----------------------------
# Query fallback helpers (PATCH)
# -----------------------------
STOP_WORDS = [
    "관광특구",
    "일대",
    "전통시장",
    "신경제 중심지",
    "호수단길",
    "관광",
    "특구",
]


def _strip_parentheses(s: str) -> str:
    # remove (...) and [...] and {...}
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"\[[^\]]*\]", " ", s)
    s = re.sub(r"\{[^}]*\}", " ", s)
    return s


def _normalize_query_text(s: str) -> str:
    s = (s or "").strip()
    s = _strip_parentheses(s)
    s = s.replace("·", " ").replace("ㆍ", " ").replace("・", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def build_fallback_queries(name: str, prefix: str = "서울 ") -> List[str]:
    """
    Generate multiple keyword queries to improve geocoding hit rate.
    """
    base = _normalize_query_text(name)
    if not base:
        return []

    queries: List[str] = []

    # 1) original (with prefix)
    queries.append(f"{prefix}{base}".strip())

    # 2) remove stop words
    alt = base
    for w in STOP_WORDS:
        alt = alt.replace(w, " ")
    alt = re.sub(r"\s+", " ", alt).strip()
    if alt and alt != base:
        queries.append(f"{prefix}{alt}".strip())

    # 3) try each chunk split by space (first token)
    first = (alt or base).split(" ")[0].strip()
    if first and first not in {base, alt}:
        queries.append(f"{prefix}{first}".strip())

    # 4) try chunks split by former middle-dot separators (best effort)
    # (after normalization, middle dot becomes space; we approximate by taking first 2 tokens too)
    tokens = (alt or base).split()
    if len(tokens) >= 2:
        queries.append(f"{prefix}{tokens[0]} {tokens[1]}".strip())
        queries.append(f"{prefix}{tokens[1]}".strip())

    # 5) remove common suffixes/patterns that can block matching
    suffix_clean = re.sub(r"(길|거리|공원|운동장)\s*$", "", (alt or base)).strip()
    if suffix_clean and suffix_clean not in {base, alt}:
        queries.append(f"{prefix}{suffix_clean}".strip())

    # keep unique, non-empty (preserve order)
    out: List[str] = []
    for q in queries:
        q = re.sub(r"\s+", " ", (q or "")).strip()
        if q and q not in out:
            out.append(q)
    return out


# -----------------------------
# XLSX parsing
# -----------------------------
@dataclass(frozen=True)
class Row:
    code: str
    name: str
    lat: Optional[float]
    lng: Optional[float]
    extra: Dict[str, Any]


CODE_COL_CANDIDATES = [
    "장소코드", "장소 코드", "코드", "AREA_CD", "AREA CODE", "place_code", "PLACE_CODE"
]
NAME_COL_CANDIDATES = [
    "장소명", "장소 명", "명칭", "AREA_NM", "AREA NAME", "place_name", "PLACE_NAME"
]
LAT_COL_CANDIDATES = ["위도", "LAT", "Latitude", "Y", "Y좌표", "Y 좌표"]
LNG_COL_CANDIDATES = ["경도", "LNG", "Longitude", "X", "X좌표", "X 좌표"]


def read_xlsx_rows(path: str, sheet: Optional[str] = None) -> List[Row]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        raise ValueError("XLSX has no header row.")

    headers = [str(h).strip() if h is not None else "" for h in header]
    colmap = guess_columns(headers)

    out: List[Row] = []
    for r in rows_iter:
        if not r:
            continue

        code = _get_cell_str(r, colmap["code_idx"])
        name = _get_cell_str(r, colmap["name_idx"])

        if not code and not name:
            continue

        lat = _get_cell_float(r, colmap.get("lat_idx"))
        lng = _get_cell_float(r, colmap.get("lng_idx"))

        extra: Dict[str, Any] = {}
        for keep in ["분류", "구분", "CATEGORY", "category", "장소구분", "서비스구분"]:
            if keep in colmap["name_to_idx"]:
                extra[keep] = _get_cell_str(r, colmap["name_to_idx"][keep])

        out.append(
            Row(
                code=code.strip() if code else "",
                name=normalize_name(name),
                lat=lat,
                lng=lng,
                extra=extra,
            )
        )

    out = [x for x in out if x.name]
    return out


def guess_columns(headers: List[str]) -> Dict[str, Any]:
    name_to_idx = {h: i for i, h in enumerate(headers) if h}

    def find_idx(candidates: List[str]) -> Optional[int]:
        for c in candidates:
            for h, i in name_to_idx.items():
                if h == c:
                    return i
        for c in candidates:
            for h, i in name_to_idx.items():
                if c.replace(" ", "").lower() == h.replace(" ", "").lower():
                    return i
        return None

    code_idx = find_idx(CODE_COL_CANDIDATES)
    name_idx = find_idx(NAME_COL_CANDIDATES)
    lat_idx = find_idx(LAT_COL_CANDIDATES)
    lng_idx = find_idx(LNG_COL_CANDIDATES)

    if code_idx is None:
        code_idx = -1
    if name_idx is None:
        raise ValueError(
            "Could not find a 장소명/AREA_NM column in XLSX header. "
            f"Headers: {headers}"
        )

    return {
        "code_idx": code_idx,
        "name_idx": name_idx,
        "lat_idx": lat_idx,
        "lng_idx": lng_idx,
        "name_to_idx": name_to_idx,
    }


# -----------------------------
# Build seed JSON
# -----------------------------
def build_seed(
    rows: List[Row],
    *,
    geocode: bool,
    kakao_key: Optional[str],
    cache_path: str,
    rate_limit_s: float,
    query_prefix: str = "서울 ",
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Returns (seed_items, missing_names)
    """
    cache = load_cache(cache_path)

    missing: List[str] = []
    items: List[Dict[str, Any]] = []

    used_codes = set()
    auto_counter = 1

    for r in rows:
        code = (r.code or "").strip()
        if not code:
            code = f"Z{auto_counter:03d}"
            auto_counter += 1

        while code in used_codes:
            code = f"{code}_DUP"
        used_codes.add(code)

        lat, lng = r.lat, r.lng

        if (lat is None or lng is None or (lat == 0.0 and lng == 0.0)) and geocode:
            key = r.name
            if key in cache:
                lat, lng = cache[key]["lat"], cache[key]["lng"]
            else:
                if not kakao_key:
                    raise RuntimeError("KAKAO_REST_API_KEY is required for --geocode.")

                found: Optional[Tuple[float, float]] = None
                tried = build_fallback_queries(r.name, prefix=query_prefix)

                for q in tried:
                    try:
                        found = kakao_keyword_geocode(query=q, rest_api_key=kakao_key)
                    except Exception as e:
                        found = None
                        print(f"[WARN] Geocode failed for '{r.name}' (query='{q}'): {e}", file=sys.stderr)

                    if found:
                        break

                if found:
                    lat, lng = found
                    cache[key] = {"lat": lat, "lng": lng}
                    save_cache(cache_path, cache)
                else:
                    missing.append(r.name)

                # rate-limit per name (not per query) to keep it simple/safer
                time.sleep(rate_limit_s)

        if lat is None or lng is None:
            missing.append(r.name)
            lat, lng = 0.0, 0.0

        items.append(
            {
                "code": code,
                "name": r.name,
                "lat": float(lat),
                "lng": float(lng),
                **({"extra": r.extra} if r.extra else {}),
            }
        )

    missing = sorted(list(set(missing)))
    return items, missing


def normalize_name(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s


# -----------------------------
# Cache helpers
# -----------------------------
def load_cache(path: str) -> Dict[str, Any]:
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f) or {}
    except Exception:
        return {}


def save_cache(path: str, cache: Dict[str, Any]) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


# -----------------------------
# Utils
# -----------------------------
def _get_cell_str(row: Tuple[Any, ...], idx: int) -> str:
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    v = row[idx]
    return "" if v is None else str(v).strip()


def _get_cell_float(row: Tuple[Any, ...], idx: Optional[int]) -> Optional[float]:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return _safe_float(row[idx], allow_none=True)


def _safe_float(v: Any, allow_none: bool = False) -> Optional[float]:
    if v is None:
        return None if allow_none else 0.0
    try:
        return float(v)
    except Exception:
        return None if allow_none else 0.0


# -----------------------------
# CLI
# -----------------------------
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Path to XLSX (120/82 list).")
    ap.add_argument("--output", required=True, help="Output JSON path (zones_seed.json).")
    ap.add_argument("--sheet", default=None, help="Excel sheet name (optional).")
    ap.add_argument(
        "--geocode",
        action="store_true",
        help="Fill missing lat/lng via Kakao keyword search (needs KAKAO_REST_API_KEY).",
    )
    ap.add_argument(
        "--cache",
        default="scripts/.cache/kakao_geocode_cache.json",
        help="Geocode cache path.",
    )
    ap.add_argument(
        "--rate-limit",
        type=float,
        default=0.25,
        help="Sleep seconds between Kakao calls (avoid rate limit).",
    )
    ap.add_argument(
        "--query-prefix",
        default="서울 ",
        help='Prefix for Kakao keyword query (default "서울 ").',
    )
    args = ap.parse_args()

    kakao_key = os.getenv("KAKAO_REST_API_KEY")

    rows = read_xlsx_rows(args.input, sheet=args.sheet)
    items, missing = build_seed(
        rows,
        geocode=args.geocode,
        kakao_key=kakao_key,
        cache_path=args.cache,
        rate_limit_s=args.rate_limit,
        query_prefix=args.query_prefix,
    )

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)

    print(f"[OK] Wrote {len(items)} zones -> {args.output}")
    if missing:
        print(f"[WARN] Missing coordinates for {len(missing)} zones. Fix or rerun geocode.")
        for n in missing[:30]:
            print(f"  - {n}")
        if len(missing) > 30:
            print("  ... (truncated)")


if __name__ == "__main__":
    main()
